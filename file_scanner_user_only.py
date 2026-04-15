#!/usr/bin/env python3
"""
File Scanner Agent - User Files Only
Scans only user-created content directories for files not accessed in over 6 months (180 days).
"""

import os
import time
from datetime import datetime, timedelta
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import sys

# Only scan these user content directories
USER_DIRECTORIES = [
    'Desktop',
    'Documents',
    'Downloads',
    'Pictures',
    'Movies',
    'Music',
    'Projects',
    'Development',
    'Code',
    'Work',
    'Bob',  # Include Bob directory
]

# Directories to exclude even within user directories
EXCLUDED_SUBDIRS = [
    '.Trash',
    'Library',
    '.cache',
    '.npm',
    '.vscode',
    '.config',
    'node_modules',
    '__pycache__',
    '.git',
    '.DS_Store',
    'venv',
    'env',
    '.idea',
    '.pytest_cache',
]

# File extensions that are critical for OS or development
EXCLUDED_EXTENSIONS = [
    '.dylib', '.framework', '.kext', '.plist', '.app',
    '.prefPane', '.bundle', '.plugin', '.component'
]

def is_excluded_path(file_path):
    """Check if the file path should be excluded."""
    file_path_str = str(file_path)
    
    # Check for excluded subdirectories
    path_parts = file_path_str.split(os.sep)
    for part in path_parts:
        if part in EXCLUDED_SUBDIRS:
            return True
    
    # Check for critical extensions
    if any(file_path_str.endswith(ext) for ext in EXCLUDED_EXTENSIONS):
        return True
    
    return False

def get_file_info(file_path):
    """Get file information including access time, modification time, and size."""
    try:
        stat_info = os.stat(file_path)
        
        # Get file size in MB
        size_mb = stat_info.st_size / (1024 * 1024)
        
        # Skip files with 0 MB
        if size_mb == 0:
            return None
        
        # Get last access time
        last_access = datetime.fromtimestamp(stat_info.st_atime)
        
        # Get last modification time
        last_modified = datetime.fromtimestamp(stat_info.st_mtime)
        
        # Get file extension
        file_extension = os.path.splitext(file_path)[1] or 'No extension'
        
        return {
            'file_path': str(file_path),
            'last_access': last_access,
            'last_modified': last_modified,
            'size_mb': round(size_mb, 2),
            'file_type': file_extension
        }
    except (OSError, PermissionError, FileNotFoundError) as e:
        # Skip files we can't access
        return None

def scan_user_directories(home_dir, cutoff_date):
    """
    Scan only user content directories for old files.
    
    Args:
        home_dir: User's home directory path
        cutoff_date: Files not accessed since this date will be included
    
    Returns:
        List of file information dictionaries
    """
    old_files = []
    scanned_count = 0
    
    print(f"Scanning user directories in: {home_dir}")
    print(f"Looking for files not accessed since: {cutoff_date.strftime('%Y-%m-%d')}")
    print("\nUser directories to scan:")
    
    # Check which directories exist
    existing_dirs = []
    for dir_name in USER_DIRECTORIES:
        dir_path = os.path.join(home_dir, dir_name)
        if os.path.exists(dir_path) and os.path.isdir(dir_path):
            existing_dirs.append(dir_path)
            print(f"  ✓ {dir_name}")
        else:
            print(f"  ✗ {dir_name} (not found)")
    
    print(f"\nScanning {len(existing_dirs)} directories...\n")
    
    for dir_path in existing_dirs:
        dir_name = os.path.basename(dir_path)
        print(f"Scanning {dir_name}...")
        
        for root, dirs, files in os.walk(dir_path):
            # Filter out excluded directories
            dirs[:] = [d for d in dirs if d not in EXCLUDED_SUBDIRS]
            
            for filename in files:
                file_path = os.path.join(root, filename)
                
                # Skip excluded paths
                if is_excluded_path(file_path):
                    continue
                
                scanned_count += 1
                if scanned_count % 500 == 0:
                    print(f"  Scanned {scanned_count} files, found {len(old_files)} old files...")
                
                file_info = get_file_info(file_path)
                
                if file_info and file_info['last_access'] < cutoff_date:
                    old_files.append(file_info)
    
    print(f"\nScan complete! Scanned {scanned_count} files.")
    print(f"Found {len(old_files)} files not accessed in over 6 months.\n")
    
    return old_files

def create_excel_report(data, output_path):
    """Create a formatted Excel report from the file data."""
    if not data:
        print("No data to write to Excel report.")
        return
    
    # Create DataFrame
    df = pd.DataFrame(data)
    
    # Sort by size (largest first)
    df = df.sort_values('size_mb', ascending=False)
    
    # Create Excel writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Old Files', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Old Files']
        
        # Format header row
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Adjust column widths
        column_widths = {
            'A': 80,  # File path
            'B': 20,  # Last access date
            'C': 20,  # Last modified date
            'D': 15,  # Size MB
            'E': 20   # File type
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # Format date columns
        for row in range(2, len(df) + 2):
            worksheet[f'B{row}'].number_format = 'YYYY-MM-DD HH:MM:SS'
            worksheet[f'C{row}'].number_format = 'YYYY-MM-DD HH:MM:SS'
            worksheet[f'D{row}'].number_format = '#,##0.00'
        
        # Add summary at the top
        worksheet.insert_rows(1, 3)
        worksheet['A1'] = 'File Inventory Report - User Files Not Accessed in Over 6 Months'
        worksheet['A1'].font = Font(bold=True, size=14)
        worksheet['A2'] = f'Report Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        worksheet['A3'] = f'Total Files Found: {len(df)}'
        
    print(f"Excel report created: {output_path}")
    print(f"Total files in report: {len(df)}")
    
    # Print summary statistics
    total_size_gb = df['size_mb'].sum() / 1024
    print(f"Total size of old files: {total_size_gb:.2f} GB")
    print(f"\nTop 5 largest files:")
    for idx, row in df.head(5).iterrows():
        print(f"  {row['size_mb']:.2f} MB - {row['file_path']}")

def main():
    """Main function to run the file scanner."""
    # Calculate cutoff date (6 months ago / 180 days)
    cutoff_date = datetime.now() - timedelta(days=180)
    
    # Get user's home directory
    home_dir = str(Path.home())
    
    # Create reports directory
    reports_dir = os.path.join(home_dir, 'Bob', 'Digital Decluttering agent', 'reports')
    os.makedirs(reports_dir, exist_ok=True)
    
    # Generate report filename with current date
    report_date = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    report_filename = f'user_files_report_{report_date}.xlsx'
    report_path = os.path.join(reports_dir, report_filename)
    
    print("=" * 70)
    print("FILE INVENTORY SCANNER - USER FILES ONLY")
    print("=" * 70)
    print(f"Scanning for files not accessed since: {cutoff_date.strftime('%Y-%m-%d')}")
    print(f"Home directory: {home_dir}")
    print(f"Report will be saved to: {report_path}")
    print("=" * 70)
    print()
    
    # Scan user directories only
    old_files = scan_user_directories(home_dir, cutoff_date)
    
    # Create Excel report
    if old_files:
        create_excel_report(old_files, report_path)
        print(f"\n✓ Report successfully created at: {report_path}")
    else:
        print("\n✓ No files found matching the criteria.")
    
    print("\nDone!")

if __name__ == '__main__':
    main()

# Made with Bob
